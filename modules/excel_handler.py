from openpyxl import load_workbook


class ExcelHandler:

    def __init__(self, file_path):

        self.workbook = load_workbook(file_path)
        self.sheet = self.workbook["Interns"]

    def get_phone_numbers(self):

        numbers = []

        for row in range(3, self.sheet.max_row + 1):

            phone = self.sheet[f"C{row}"].value

            if phone:
                numbers.append((row, str(phone)))

        return numbers

    def write_offer(self, row, column, value):

        self.sheet[f"{column}{row}"] = value

    def save(self):

        self.workbook.save(r"https://airtelaf-my.sharepoint.com/:x:/g/personal/13400403_airtel_africa/IQAnROVW_apXSJ2A0axNmhZiAcmSDh1qT8AUatkIzirW66Y?e=eJOza1&isSPOFile=1&ovuser=16c73727-979c-4d82-b3a7-eb6a2fddfe57%2C23245518%40airtel.africa&wdExp=TEAMS-TREATMENT&web=1&TeamsCID=2c21f2f2-9c46-49d3-b7a8-f0af79d561a1&clickparams=eyJBcHBOYW1lIjoiVGVhbXMtRGVza3RvcCIsIkFwcFZlcnNpb24iOiI0OS8yNjA1MjkwNjEyMSJ9&linkOpenTime=1782796862307")